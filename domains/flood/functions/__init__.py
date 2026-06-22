from __future__ import annotations

import json
import re
import subprocess
from functools import cached_property
from pathlib import Path
from typing import Any

import openpyxl

from oag.ontology.registry import FunctionRegistry
from oag.ontology.repository import ObjectRepository
from oag.ontology.schema import Ontology


DOMAIN_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = DOMAIN_DIR.parents[1]
DATA_DIR = PROJECT_DIR / "珊瑚河数据"
GENERATED_DIR = DOMAIN_DIR / "data" / "generated"


SCENARIO_CODE_ALIASES = {
    "45050092shh0001": "45050092hsfx0001",
    "45050092shh0002": "45050092hsfx0002",
    "45050092shh0003": "45050092hsfx0003",
    "45050092shh0004": "45050092hsfx0004",
    "45050092shh0005": "45050092hsfx0005",
}

MAPPABLE_OBJECTS = {
    "River": {
        "source": DATA_DIR / "1.流域边界/珊瑚河流域范围.shp",
        "label": "珊瑚河",
        "role": "base",
        "style": {"type": "line", "color": "#111827", "weight": 1},
    },
    "County": {
        "source": DATA_DIR / "2.县界/珊瑚河县界.shp",
        "label": "县级行政区",
        "role": "base",
        "style": {"type": "line", "color": "#64748b", "weight": 1},
    },
    "Road": {
        "source": DATA_DIR / "5.路网/公路-线.shp",
        "label": "道路",
        "role": "base",
        "style": {"type": "line", "color": "#6b7280", "weight": 2},
    },
    "Reservoir": {
        "source": DATA_DIR / "3.水利工程/水库.shp",
        "label": "水库",
        "role": "asset",
        "style": {"type": "circle", "color": "#2563eb", "radius": 5, "stroke": "#ffffff"},
    },
    "Sluice": {
        "source": DATA_DIR / "3.水利工程/水闸.shp",
        "label": "水闸",
        "role": "asset",
        "style": {"type": "circle", "color": "#0891b2", "radius": 5, "stroke": "#ffffff"},
    },
    "Bridge": {
        "source": DATA_DIR / "3.水利工程/桥梁.shp",
        "label": "桥梁",
        "role": "asset",
        "style": {"type": "circle", "color": "#111827", "radius": 5, "stroke": "#ffffff"},
    },
    "Facility": {
        "label": "重要设施",
        "role": "asset",
        "style": {"type": "circle", "color": "#dc2626", "radius": 5, "stroke": "#ffffff"},
    },
    "HydraulicStructure": {
        "label": "水利工程设施",
        "role": "asset",
        "style": {"type": "circle", "color": "#0f766e", "radius": 5, "stroke": "#ffffff"},
    },
    "Place": {
        "source": DATA_DIR / "8.避洪转移/安置点.shp",
        "label": "安置地点",
        "role": "evacuation",
        "style": {"type": "circle", "color": "#16a34a", "radius": 5, "stroke": "#ffffff"},
    },
    "Transfer": {
        "source": DATA_DIR / "8.避洪转移/转移单元.shp",
        "label": "转移安排",
        "role": "evacuation",
        "style": {"type": "circle", "color": "#f97316", "radius": 5, "stroke": "#ffffff"},
    },
    "Route": {
        "source": DATA_DIR / "8.避洪转移/转移路线.shp",
        "label": "路线",
        "role": "evacuation",
        "style": {"type": "line", "color": "#ef4444", "weight": 3},
    },
    "Cell": {
        "label": "洪水计算单元",
        "role": "scenario",
        "style": {"type": "fill", "fillColor": "#4292c6", "fillOpacity": 0.35, "color": "#4292c6", "weight": 0.5},
    },
}


class FloodRepository:
    def __init__(self):
        self._row_cache: dict[str, list[dict]] = {}
        self._object_loaders = {
            "River": self._load_rivers,
            "County": self._load_counties,
            "Town": self._load_towns,
            "Reservoir": self._load_reservoirs,
            "Sluice": self._load_sluices,
            "Bridge": self._load_bridges,
            "Facility": self._load_facilities,
            "HydraulicStructure": self._load_hydraulic_structures,
            "Road": self._load_roads,
            "Place": self._load_places,
            "Transfer": self._load_transfers,
            "Route": self._load_routes,
            "Scenario": self._load_scenarios,
            "Cell": self._load_cells,
            "Impact": self._load_impacts,
            "Hydrology": self._load_hydrology,
        }

    def query(self, object_type: str, filters: dict[str, Any] | None = None,
              limit: int | None = None, order_by: str | None = None,
              offset: int | None = None) -> list[dict]:
        if object_type == "Cell":
            return self._query_cells(filters, limit, order_by, offset)
        rows = [dict(row) for row in self._rows(object_type)]
        rows = _apply_filters(rows, filters)
        rows = _apply_order(rows, order_by)
        return _apply_window(rows, limit, offset)

    def count(self, object_type: str, filters: dict[str, Any] | None = None) -> int:
        if object_type == "Cell":
            return self._count_cells(filters)
        return len(self.query(object_type, filters))

    def query_by_id(self, object_type: str, id_value: Any) -> dict | None:
        id_field = _id_field(object_type)
        rows = self.query(object_type, {id_field: id_value}, limit=1)
        return rows[0] if rows else None

    def search_text(self, keyword: str, object_types: list[str] | None = None,
                    limit: int = 20) -> list[dict]:
        if not keyword:
            return []
        results = []
        searchable_types = object_types or [
            item for item in self._object_loaders if item != "Cell"
        ]
        for object_type in searchable_types:
            for row in self._rows(object_type):
                matched = [
                    key for key, value in row.items()
                    if isinstance(value, str) and keyword in value
                ]
                if not matched:
                    continue
                result = dict(row)
                result["_object_type"] = object_type
                result["_matched_field"] = ", ".join(matched)
                results.append(result)
                if len(results) >= limit:
                    return results
        return results

    def _rows(self, object_type: str) -> list[dict]:
        if object_type in self._row_cache:
            return self._row_cache[object_type]
        loader = self._object_loaders.get(object_type)
        rows = loader() if loader else []
        if object_type != "Cell":
            self._row_cache[object_type] = rows
        return rows

    def _query_cells(self, filters: dict[str, Any] | None = None,
                     limit: int | None = None, order_by: str | None = None,
                     offset: int | None = None) -> list[dict]:
        scenario = _find_scenario(
            self,
            (filters or {}).get("scenario_id", ""),
            int((filters or {}).get("return_period_year", 0) or 0),
        )
        scenarios = [scenario] if scenario else self.scenarios[:1]
        rows = []
        for item in scenarios:
            source = _scenario_path(item)
            for i, feature in enumerate(_features(source), 1):
                rows.append(_cell_record(item["scenario_id"], i, feature, source))
        rows = _apply_filters(rows, filters)
        rows = _apply_order(rows, order_by)
        return _apply_window(rows, limit, offset)

    def _count_cells(self, filters: dict[str, Any] | None = None) -> int:
        if not filters or set(filters) <= {"scenario_id", "return_period_year"}:
            scenario = _find_scenario(
                self,
                (filters or {}).get("scenario_id", ""),
                int((filters or {}).get("return_period_year", 0) or 0),
            )
            scenarios = [scenario] if scenario else self.scenarios
            return sum(_ogr_metadata(_scenario_path(item)).get("feature_count", 0) for item in scenarios if item)
        return len(self.query("Cell", filters))

    @cached_property
    def scenarios(self) -> list[dict]:
        return self._read_scenarios()

    @cached_property
    def impacts(self) -> list[dict]:
        return self._read_impacts()

    @cached_property
    def hydrology(self) -> list[dict]:
        return self._read_hydrology()

    @cached_property
    def towns(self) -> list[dict]:
        return self._read_towns()

    def _load_rivers(self) -> list[dict]:
        feature = _features(DATA_DIR / "1.流域边界/珊瑚河流域范围.shp")
        geo = _geometry_fields(feature[0]) if feature else {}
        return [{
            "river_id": "shanhu",
            "name": "珊瑚河",
            "basin_name": "珊瑚河流域",
            "crs": "EPSG:2434",
            **geo,
            "data_path": _rel(DATA_DIR / "1.流域边界/珊瑚河流域范围.shp"),
        }]

    def _load_counties(self) -> list[dict]:
        rows = []
        for i, feature in enumerate(_features(DATA_DIR / "2.县界/珊瑚河县界.shp"), 1):
            props = feature.get("properties") or {}
            adcode = _code(_first_non_empty(props, "adcode", "ADCODE")) or f"county_{i}"
            rows.append({
                "county_id": adcode,
                "river_id": "shanhu",
                "name": _first_non_empty(props, "Name", "name") or "钟山县",
                "adcode": adcode,
                **_geometry_fields(feature),
                "data_path": _rel(DATA_DIR / "2.县界/珊瑚河县界.shp"),
            })
        return rows

    def _load_towns(self) -> list[dict]:
        return self.towns

    def _load_reservoirs(self) -> list[dict]:
        return [
            _reservoir_record(i, feature, DATA_DIR / "3.水利工程/水库.shp")
            for i, feature in enumerate(_features(DATA_DIR / "3.水利工程/水库.shp"), 1)
        ]

    def _load_sluices(self) -> list[dict]:
        return [
            _sluice_record(i, feature, DATA_DIR / "3.水利工程/水闸.shp")
            for i, feature in enumerate(_features(DATA_DIR / "3.水利工程/水闸.shp"), 1)
        ]

    def _load_bridges(self) -> list[dict]:
        return [
            _bridge_record(i, feature, DATA_DIR / "3.水利工程/桥梁.shp")
            for i, feature in enumerate(_features(DATA_DIR / "3.水利工程/桥梁.shp"), 1)
        ]

    def _load_facilities(self) -> list[dict]:
        specs = [
            ("hospital", DATA_DIR / "4.重要设施/医院.shp"),
            ("school", DATA_DIR / "4.重要设施/学校.shp"),
            ("government", DATA_DIR / "4.重要设施/政府.shp"),
        ]
        rows = []
        for facility_type, path in specs:
            for i, feature in enumerate(_features(path), 1):
                rows.append(_facility_record(facility_type, i, feature, path))
        return rows

    def _load_hydraulic_structures(self) -> list[dict]:
        specs = [
            ("levee", DATA_DIR / "3.水利工程/堤防.shp"),
            ("pump_station", DATA_DIR / "3.水利工程/泵站.shp"),
            ("spillway", DATA_DIR / "3.水利工程/溢流建筑物.shp"),
        ]
        rows = []
        for structure_type, path in specs:
            for i, feature in enumerate(_features(path), 1):
                rows.append(_hydraulic_structure_record(structure_type, i, feature, path))
        return rows

    def _load_roads(self) -> list[dict]:
        return [
            _road_record(i, feature, DATA_DIR / "5.路网/公路-线.shp")
            for i, feature in enumerate(_features(DATA_DIR / "5.路网/公路-线.shp"), 1)
        ]

    def _load_places(self) -> list[dict]:
        rows = []
        for i, feature in enumerate(_features(DATA_DIR / "8.避洪转移/安置点.shp"), 1):
            rows.append(_place_record("shelter", i, feature, DATA_DIR / "8.避洪转移/安置点.shp"))
        for i, feature in enumerate(_features(DATA_DIR / "8.避洪转移/就地单元.shp"), 1):
            rows.append(_place_record("in_place", i, feature, DATA_DIR / "8.避洪转移/就地单元.shp"))
        return rows

    def _load_transfers(self) -> list[dict]:
        transfer_by_id = {}
        for i, feature in enumerate(_features(DATA_DIR / "8.避洪转移/转移单元.shp"), 1):
            row = _transfer_record(i, feature, DATA_DIR / "8.避洪转移/转移单元.shp")
            transfer_by_id[row["transfer_id"]] = row
        for route in self._load_routes():
            transfer_id = route.get("transfer_id")
            if transfer_id in transfer_by_id:
                transfer_by_id[transfer_id]["route_id"] = route["route_id"]
                transfer_by_id[transfer_id]["place_id"] = route.get("place_id", "")
        return list(transfer_by_id.values())

    def _load_routes(self) -> list[dict]:
        return [
            _route_record(i, feature, DATA_DIR / "8.避洪转移/转移路线.shp")
            for i, feature in enumerate(_features(DATA_DIR / "8.避洪转移/转移路线.shp"), 1)
        ]

    def _load_scenarios(self) -> list[dict]:
        return self.scenarios

    def _load_cells(self) -> list[dict]:
        rows = []
        for scenario in self.scenarios:
            source = _scenario_path(scenario)
            for i, feature in enumerate(_features(source), 1):
                rows.append(_cell_record(scenario["scenario_id"], i, feature, source))
        return rows

    def _load_impacts(self) -> list[dict]:
        return self.impacts

    def _load_hydrology(self) -> list[dict]:
        return self.hydrology

    def _read_scenarios(self) -> list[dict]:
        path = DATA_DIR / "6.淹没图层/方案说明内容_河流.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            river_name, _, scenario_id, name, description, summary = row[:6]
            if not scenario_id:
                continue
            rows.append({
                "scenario_id": str(scenario_id),
                "river_id": "shanhu",
                "name": str(name or ""),
                "return_period_year": _return_period(name or summary or description),
                "summary": str(summary or ""),
                "description": str(description or ""),
                "data_path": _rel(DATA_DIR / f"6.淹没图层/45050092_珊瑚河/{scenario_id}.shp"),
            })
        wb.close()
        return rows

    def _read_impacts(self) -> list[dict]:
        path = DATA_DIR / "7.灾前&灾损数据/45050092_洪水影响分析与损失评估结果表.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            raw_id = str(row[0])
            scenario_id = SCENARIO_CODE_ALIASES.get(raw_id, raw_id)
            rows.append({
                "impact_id": f"impact_{scenario_id}",
                "scenario_id": scenario_id,
                "inundated_area_km2": _float(row[1]),
                "inundated_building_10k_m2": _float(row[2]),
                "inundated_farmland_ha": _float(row[3]),
                "inundated_road_km": _float(row[4]),
                "affected_population_10k": _float(row[5]),
                "affected_gdp_10k_cny": _float(row[6]),
                "direct_loss_10k_cny": _float(row[7]),
            })
        wb.close()
        return rows

    def _read_hydrology(self) -> list[dict]:
        path = DATA_DIR / "9.水文计算结果/珊瑚河水文分析结果.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        periods = [100, 50, 20, 10, 5]
        scenario_by_period = {
            row["return_period_year"]: row["scenario_id"]
            for row in self.scenarios
            if row.get("return_period_year")
        }
        for r in range(4, 7):
            duration = _duration_hours(ws.cell(r, 1).value)
            for idx, period in enumerate(periods, start=5):
                rows.append({
                    "hydrology_id": f"rain_{duration:g}h_{period}a",
                    "river_id": "shanhu",
                    "scenario_id": scenario_by_period.get(period, ""),
                    "return_period_year": period,
                    "duration_h": duration,
                    "design_rainfall_mm": _float(ws.cell(r, idx).value),
                    "source": "珊瑚河水文分析结果.xlsx: 设计暴雨",
                })
        peak_values = [ws.cell(12, col).value for col in range(3, 8)]
        discharge_values = [ws.cell(13, col).value for col in range(3, 8)]
        for period, peak, discharge in zip(periods, peak_values, discharge_values):
            rows.append({
                "hydrology_id": f"flood_peak_{period}a",
                "river_id": "shanhu",
                "scenario_id": scenario_by_period.get(period, ""),
                "return_period_year": period,
                "peak_inflow_m3s": _float(peak),
                "max_discharge_m3s": _float(discharge),
                "source": "珊瑚河水文分析结果.xlsx: 设计洪水",
            })
        wb.close()
        return rows

    def _read_towns(self) -> list[dict]:
        path = DATA_DIR / "7.灾前&灾损数据/珊瑚河灾前信息.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0])
            rows.append({
                "town_id": _slug_id(name),
                "name": name,
                "county_id": "451122",
                "population": int(_float(row[2])),
                "gdp_10k_cny": _float(row[1]),
                "agriculture_10k_cny": _float(row[3]),
                "housing_property_10k_cny": _float(row[4]),
                "transport_industry_10k_cny": _float(row[5]),
                "commerce_10k_cny": _float(row[6]),
            })
        wb.close()
        return rows


def register(registry: FunctionRegistry, repository: ObjectRepository,
             ontology: Ontology):
    resolver = FloodRepository()
    registry.register_resolver("flood_repository", resolver)

    registry.register("list_scenarios", lambda: resolver.scenarios,
                      ontology.functions["list_scenarios"])
    registry.register("get_scenario_summary", lambda scenario_id="", return_period_year=0: _scenario_summary(
        resolver, scenario_id, return_period_year,
    ), ontology.functions["get_scenario_summary"])
    registry.register("analyze_risks", _not_wired("analyze_risks"),
                      ontology.functions["analyze_risks"])
    registry.register("list_mappable_objects", lambda object_type="": _list_mappable_objects(
        resolver, object_type,
    ), ontology.functions["list_mappable_objects"])
    registry.register("export_objects_geojson", lambda object_type, filters=None, simplify_tolerance=0, force=False: _export_objects_geojson(
        resolver, object_type, filters or {}, simplify_tolerance, force,
    ), ontology.functions["export_objects_geojson"])
    registry.register("plan_response", _not_wired("plan_response"),
                      ontology.functions["plan_response"])
    registry.register("generate_brief", _not_wired("generate_brief"),
                      ontology.functions["generate_brief"])


def _list_mappable_objects(resolver: FloodRepository, object_type: str = "") -> list[dict]:
    object_types = [object_type] if object_type else list(MAPPABLE_OBJECTS)
    rows = []
    for item in object_types:
        spec = MAPPABLE_OBJECTS.get(item)
        if not spec:
            continue
        if item == "Cell":
            count = sum(_ogr_metadata(_scenario_path(row)).get("feature_count", 0) for row in resolver.scenarios)
            geometry_type = "Polygon"
        else:
            sample = resolver.query(item, limit=1)
            count = resolver.count(item)
            geometry_type = sample[0].get("geometry_type", "") if sample else ""
        rows.append({
            "object_type": item,
            "label": spec.get("label", item),
            "role": spec.get("role", ""),
            "feature_count": count,
            "geometry_type": geometry_type,
            "map_crs": "EPSG:4326",
            "default_style": spec.get("style", {}),
        })
    return rows


def _scenario_summary(resolver: FloodRepository, scenario_id: str = "",
                      return_period_year: int = 0) -> dict:
    scenario = _find_scenario(resolver, scenario_id, return_period_year)
    if not scenario:
        return {"error": "scenario not found", "scenario_id": scenario_id,
                "return_period_year": return_period_year}
    sid = scenario["scenario_id"]
    return {
        "scenario": scenario,
        "impact": next((row for row in resolver.impacts if row["scenario_id"] == sid), None),
        "hydrology": [row for row in resolver.hydrology if row.get("scenario_id") == sid],
        "mappable": {
            "object_type": "Cell",
            "filters": {"scenario_id": sid},
            "export_tool": "export_objects_geojson",
        },
    }


def _export_objects_geojson(resolver: FloodRepository, object_type: str,
                            filters: dict[str, Any] | None = None,
                            simplify_tolerance: float = 0,
                            force: bool = False) -> dict:
    filters = filters or {}
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    key = _export_key(object_type, filters)
    suffix = f"_s{simplify_tolerance:g}" if simplify_tolerance else ""
    target = GENERATED_DIR / f"{key}{suffix}.geojson"
    if target.exists() and not force:
        return _geojson_result(object_type, filters, target, cached=True)

    if object_type == "Cell":
        scenario = _find_scenario(resolver, filters.get("scenario_id", ""), int(filters.get("return_period_year", 0) or 0))
        if not scenario:
            return {"error": "scenario not found", "filters": filters}
        source = _scenario_path(scenario)
        _ogr_export(source, target, simplify_tolerance)
        return _geojson_result(object_type, {"scenario_id": scenario["scenario_id"]}, target, cached=False)

    rows = resolver.query(object_type, filters=filters)
    collection = {
        "type": "FeatureCollection",
        "name": object_type,
        "features": [_feature_from_row(row) for row in rows if row.get("geometry")],
    }
    target.write_text(json.dumps(collection, ensure_ascii=False), encoding="utf-8")
    return _geojson_result(object_type, filters, target, cached=False)


def _ogr_export(source: Path, target: Path, simplify_tolerance: float = 0):
    cmd = [
        "ogr2ogr",
        "-f", "GeoJSON",
        "-t_srs", "EPSG:4326",
        "-lco", "COORDINATE_PRECISION=7",
    ]
    if simplify_tolerance:
        cmd.extend(["-simplify", str(simplify_tolerance)])
    cmd.extend([str(target), str(source)])
    subprocess.run(cmd, check=True, capture_output=True, text=True)


def _geojson_result(object_type: str, filters: dict, target: Path, cached: bool) -> dict:
    spec = MAPPABLE_OBJECTS.get(object_type, {})
    return {
        "object_type": object_type,
        "label": spec.get("label", object_type),
        "filters": filters,
        "path": _rel(target),
        "absolute_path": str(target),
        "crs": "EPSG:4326",
        "cached": cached,
        "default_style": spec.get("style", {}),
    }


def _not_wired(name: str):
    def handler(**kwargs):
        return {
            "status": "not_implemented",
            "tool": name,
            "message": "data listing and map export are wired; spatial risk analysis is the next implementation step.",
            "args": kwargs,
        }

    return handler


def _find_scenario(resolver: FloodRepository, scenario_id: str = "",
                   return_period_year: int = 0) -> dict | None:
    if scenario_id:
        return next((row for row in resolver.scenarios if row["scenario_id"] == scenario_id), None)
    if return_period_year:
        return next((row for row in resolver.scenarios
                     if row.get("return_period_year") == int(return_period_year)), None)
    return resolver.scenarios[0] if resolver.scenarios else None


def _ogr_metadata(path: Path) -> dict:
    try:
        raw = subprocess.check_output(
            ["ogrinfo", "-json", "-so", "-al", str(path)],
            text=True,
            stderr=subprocess.STDOUT,
        )
        data = json.loads(raw)
        layer = data.get("layers", [{}])[0]
        geom = (layer.get("geometryFields") or [{}])[0]
        epsg = _epsg_from_geom(geom)
        extent = geom.get("extent") or []
        return {
            "geometry_type": geom.get("type", ""),
            "source_crs": epsg,
            "extent": extent,
            "feature_count": layer.get("featureCount", 0),
        }
    except Exception as exc:
        return {"error": str(exc)}


def _features(path: Path) -> list[dict]:
    result = subprocess.run(
        ["ogr2ogr", "-f", "GeoJSON", "/vsistdout/", "-t_srs", "EPSG:4326", str(path)],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0 or not result.stdout.strip():
        return []
    return json.loads(result.stdout).get("features", [])


def _reservoir_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_rid = _first_non_empty(props, "CD", "ennmcd", "水库_1", "水库编", "OBJECTID") or "reservoir"
    rid = f"{_code(raw_rid) or raw_rid}_{index}"
    return {
        "reservoir_id": str(rid),
        "name": _first_non_empty(props, "NAME", "ennm", "水库名") or f"水库{index}",
        "river_id": "shanhu" if _first_non_empty(props, "所在河", "RV_NAME") == "珊瑚河" else "",
        "county_id": _code(_first_non_empty(props, "行政区", "adcode")),
        "town_name": _first_non_empty(props, "乡（镇", "town_name") or "",
        "capacity_10k_m3": _float(_first_non_empty(props, "CAPACITY", "总库容")),
        "dam_top_elevation_m": _float(_first_non_empty(props, "MAX_H", "坝顶高")),
        "design_flood_standard_year": _float(_first_non_empty(props, "D_ST", "设计洪")),
        "check_flood_standard_year": _float(_first_non_empty(props, "CH_ST", "校核洪")),
        "max_discharge_m3s": _float(_first_non_empty(props, "V_MAXDIS", "最大泄")),
        "safety_status": str(_first_non_empty(props, "SAFETY", "审核标") or ""),
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _sluice_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_sid = _first_non_empty(props, "CODE", "水闸编", "OBJECTID", "序号") or "sluice"
    sid = f"{_code(raw_sid) or raw_sid}_{index}"
    river_name = _first_non_empty(props, "RV_NAME", "所在河") or ""
    return {
        "sluice_id": str(sid),
        "name": _first_non_empty(props, "NAME", "水闸名") or f"水闸{index}",
        "river_id": "shanhu" if river_name == "珊瑚河" else "",
        "county_id": _code(_first_non_empty(props, "行政区", "adcode")),
        "town_name": _first_non_empty(props, "乡（镇") or "",
        "sluice_type": _first_non_empty(props, "TYPE", "水闸类") or "",
        "gate_count": int(_float(_first_non_empty(props, "闸孔数"))),
        "design_discharge_m3s": _float(_first_non_empty(props, "分_泄_", "D_MAXDIS")),
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _bridge_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    name = _first_non_empty(props, "NAME", "ennm", "名称_nam")
    bridge_id = _first_non_empty(props, "CODE", "ennmcd", "Id") or f"bridge_{index}"
    populated = [value for value in props.values() if value not in (None, "", 0, 0.0, [])]
    return {
        "bridge_id": str(bridge_id) if str(bridge_id) != "0" else f"bridge_{index}",
        "name": name or f"桥梁{index}",
        "river_id": "shanhu" if _first_non_empty(props, "RV_NAME") == "珊瑚河" else "",
        "road_id": "",
        "length_m": _float(_first_non_empty(props, "LEN")),
        "width_m": _float(_first_non_empty(props, "WIDE")),
        "deck_elevation_m": _float(_first_non_empty(props, "EL", "B_EL")),
        "data_quality": "partial" if len(populated) > 3 else "missing",
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _facility_record(facility_type: str, index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    fid = _first_non_empty(props, "id", "OBJECTID") or f"{facility_type}_{index}"
    return {
        "facility_id": str(fid),
        "name": _first_non_empty(props, "name") or f"{_facility_type_cn(facility_type)}{index}",
        "facility_type": facility_type,
        "subtype": _first_non_empty(props, "行业小", "行业中") or "",
        "address": _first_non_empty(props, "address") or "",
        "town_name": "",
        "county_id": _code(_first_non_empty(props, "adcode")),
        "longitude": lng or _float(_first_non_empty(props, "wgs84_经", "longitude")),
        "latitude": lat or _float(_first_non_empty(props, "wgs84_纬", "latitude")),
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _hydraulic_structure_record(structure_type: str, index: int,
                                feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_sid = _first_non_empty(props, "ID", "CODE", "ennmcd", "OBJECTID") or structure_type
    sid = f"{_code(raw_sid) or raw_sid}_{index}"
    river_name = _first_non_empty(props, "RV_NAME", "LOC", "所在河") or ""
    return {
        "structure_id": str(sid),
        "name": _first_non_empty(props, "NAME", "ennm", "enname") or f"{_structure_type_cn(structure_type)}{index}",
        "structure_type": structure_type,
        "river_id": "shanhu" if river_name == "珊瑚河" else "",
        "river_name": river_name,
        "county_id": _code(_first_non_empty(props, "adcode", "行政区")),
        "location": _first_non_empty(props, "LOC") or "",
        "length_m": _float(_first_non_empty(props, "LEN", "Shape_Leng")),
        "elevation_m": _float(_first_non_empty(props, "EL", "MAX_H")),
        "flow_m3s": _float(_first_non_empty(props, "FLOW", "D_MAXDIS")),
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _road_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    road_id = _first_non_empty(props, "osm_id", "OBJECTID") or f"road_{index}"
    return {
        "road_id": str(road_id),
        "name": _first_non_empty(props, "name") or _first_non_empty(props, "ref") or f"道路{index}",
        "ref": _first_non_empty(props, "ref") or "",
        "road_class": _first_non_empty(props, "fclass") or "",
        "one_way": _first_non_empty(props, "oneway") or "",
        "bridge_flag": str(_first_non_empty(props, "bridge") or "").upper() == "T",
        "tunnel_flag": str(_first_non_empty(props, "tunnel") or "").upper() == "T",
        "length_m": _float(_first_non_empty(props, "Shape_Leng")),
        "crs": "EPSG:4546",
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _place_record(place_type: str, index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_id = _first_non_empty(props, "ID", "OBJECTID") or f"{place_type}_{index}"
    return {
        "place_id": f"{place_type}_{_code(raw_id) or raw_id}",
        "name": _first_non_empty(props, "Name") or f"{_place_type_cn(place_type)}{index}",
        "place_type": place_type,
        "town_id": _code(_first_non_empty(props, "OwnerTown_")),
        "town_name": _first_non_empty(props, "OwnerTown") or "",
        "county_id": _code(_first_non_empty(props, "OwnerQX_Co")),
        "area_m2": _float(_first_non_empty(props, "Area")),
        "capacity_person": int(_float(_first_non_empty(props, "Vol", "Population"))),
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _transfer_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    transfer_id = _code(_first_non_empty(props, "ID", "OBJECTID")) or f"transfer_{index}"
    return {
        "transfer_id": transfer_id,
        "name": _first_non_empty(props, "Name") or f"转移{index}",
        "population": int(_float(_first_non_empty(props, "Population"))),
        "town_id": _code(_first_non_empty(props, "OwnerTown_")),
        "town_name": _first_non_empty(props, "OwnerTown") or "",
        "county_id": _code(_first_non_empty(props, "OwnerQX_Co")),
        "arrive_time_window": _first_non_empty(props, "ArriveTime") or "",
        "route_id": "",
        "place_id": "",
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _route_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    route_id = _code(_first_non_empty(props, "ID", "OBJECTID")) or f"route_{index}"
    transfer_id = _code(_first_non_empty(props, "waterID", "ID"))
    place_raw_id = _code(_first_non_empty(props, "safeID"))
    return {
        "route_id": route_id,
        "name": _first_non_empty(props, "TransferNa") or f"路线{index}",
        "route_type": "transfer",
        "road_detail": _first_non_empty(props, "RoadDetail") or "",
        "transfer_id": transfer_id,
        "place_id": f"shelter_{place_raw_id}" if place_raw_id else "",
        "population": int(_float(_first_non_empty(props, "Population"))),
        "length_m": _float(_first_non_empty(props, "Shape_Leng")),
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _cell_record(scenario_id: str, index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    return {
        "cell_id": f"{scenario_id}_{_code(_first_non_empty(props, 'OBJECTID', 'FID')) or index}",
        "scenario_id": scenario_id,
        "x": _float(_first_non_empty(props, "X", "x")),
        "y": _float(_first_non_empty(props, "Y", "y")),
        "ground_elevation_m": _float(_first_non_empty(props, "Z", "z")),
        "water_level_m": _float(_first_non_empty(props, "Eta", "ETA")),
        "depth_m": _float(_first_non_empty(props, "YMSS", "depth")),
        "velocity_mps": _float(_first_non_empty(props, "HSLS", "velocity")),
        "arrival_time_h": _float(_first_non_empty(props, "DDSJ", "arrival")),
        "recession_time_h": _float(_first_non_empty(props, "XTSJ", "recession")),
        "area_m2": _float(_first_non_empty(props, "WGMJ", "Shape_Area")),
        "crs": "EPSG:4546",
        **_geometry_fields(feature),
        "data_path": _rel(path),
    }


def _transform_extent(path: Path) -> list[float]:
    raw = subprocess.check_output(
        ["ogrinfo", "-json", "-so", "-al", str(path)],
        text=True,
        stderr=subprocess.STDOUT,
    )
    data = json.loads(raw)
    geom = (data.get("layers", [{}])[0].get("geometryFields") or [{}])[0]
    return geom.get("extent") or []


def _epsg_from_geom(geom: dict) -> str:
    coord = geom.get("coordinateSystem") or {}
    ids = re.findall(r'ID\["EPSG",(\d+)\]', coord.get("wkt", ""))
    return f"EPSG:{ids[-1]}" if ids else ""


def _apply_filters(rows: list[dict], filters: dict[str, Any] | None) -> list[dict]:
    result = list(rows)
    for key, value in (filters or {}).items():
        field, op = key.split("__", 1) if "__" in key else (key, "eq")
        if op == "like":
            result = [row for row in result if str(value) in str(row.get(field, ""))]
        elif op == "ne":
            result = [row for row in result if row.get(field) != value]
        elif op == "gt":
            result = [row for row in result if row.get(field) > value]
        elif op == "gte":
            result = [row for row in result if row.get(field) >= value]
        elif op == "lt":
            result = [row for row in result if row.get(field) < value]
        elif op == "lte":
            result = [row for row in result if row.get(field) <= value]
        else:
            result = [row for row in result if row.get(field) == value]
    return result


def _apply_order(rows: list[dict], order_by: str | None) -> list[dict]:
    if not order_by:
        return rows
    reverse = order_by.startswith("-")
    field = order_by.lstrip("-")
    return sorted(rows, key=lambda row: row.get(field), reverse=reverse)


def _apply_window(rows: list[dict], limit: int | None,
                  offset: int | None) -> list[dict]:
    if offset:
        rows = rows[offset:]
    if limit:
        rows = rows[:limit]
    return rows


def _id_field(object_type: str) -> str:
    return {
        "River": "river_id",
        "County": "county_id",
        "Town": "town_id",
        "Reservoir": "reservoir_id",
        "Sluice": "sluice_id",
        "HydraulicStructure": "structure_id",
        "Road": "road_id",
        "Bridge": "bridge_id",
        "Facility": "facility_id",
        "Place": "place_id",
        "Transfer": "transfer_id",
        "Route": "route_id",
        "Scenario": "scenario_id",
        "Impact": "impact_id",
        "Hydrology": "hydrology_id",
    }.get(object_type, f"{object_type.lower()}_id")


def _rel(path: Path | str) -> str:
    path = Path(path)
    try:
        return str(path.resolve().relative_to(PROJECT_DIR.resolve()))
    except ValueError:
        return str(path)


def _float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _first_non_empty(values: dict, *keys: str) -> Any:
    for key in keys:
        value = values.get(key)
        if value not in (None, "", [], {}, "null"):
            return value
    return None


def _point_coords(feature: dict) -> tuple[float, float]:
    geom = feature.get("geometry") or {}
    coords = geom.get("coordinates") or []
    if geom.get("type") == "Point" and len(coords) >= 2:
        return float(coords[0]), float(coords[1])
    return 0.0, 0.0


def _geometry_fields(feature: dict) -> dict:
    geom = feature.get("geometry") or {}
    return {
        "geometry_type": geom.get("type", ""),
        "geometry": json.dumps(geom, ensure_ascii=False) if geom else "",
    }


def _feature_from_row(row: dict) -> dict:
    geometry = json.loads(row.get("geometry") or "{}")
    properties = {
        key: value for key, value in row.items()
        if key not in {"geometry", "geometry_type"}
    }
    return {
        "type": "Feature",
        "properties": properties,
        "geometry": geometry,
    }


def _export_key(object_type: str, filters: dict[str, Any]) -> str:
    if object_type == "Cell":
        scenario_id = filters.get("scenario_id") or f"{filters.get('return_period_year', '')}a"
        return f"{object_type.lower()}_{scenario_id}".strip("_")
    if not filters:
        return object_type.lower()
    parts = [object_type.lower()]
    for key in sorted(filters):
        value = str(filters[key]).replace("/", "_").replace(" ", "_")
        parts.append(f"{key}_{value}")
    return "__".join(parts)


def _scenario_path(scenario: dict) -> Path:
    return PROJECT_DIR / scenario["data_path"]


def _facility_type_cn(value: str) -> str:
    return {"hospital": "医院", "school": "学校", "government": "政府"}.get(value, "设施")


def _place_type_cn(value: str) -> str:
    return {"shelter": "安置点", "in_place": "就地安置点"}.get(value, "地点")


def _structure_type_cn(value: str) -> str:
    return {
        "levee": "堤防",
        "pump_station": "泵站",
        "spillway": "溢流建筑物",
    }.get(value, "水利工程")


def _return_period(value: Any) -> int:
    match = re.search(r"(\d+)\s*(?:a|年)", str(value))
    return int(match.group(1)) if match else 0


def _duration_hours(value: Any) -> float:
    match = re.search(r"(\d+(?:\.\d+)?)", str(value))
    return float(match.group(1)) if match else 0


def _slug_id(value: str) -> str:
    mapping = {
        "回龙镇": "451122104",
        "石龙镇": "451122105",
        "凤翔镇": "451122106",
        "珊瑚镇": "451122107",
        "同古镇": "451122108",
        "清塘镇": "451122111",
    }
    return mapping.get(value, value)


def _code(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    return text[:-2] if text.endswith(".0") else text


def _style(color: str, fill: bool = True, point: bool = False,
           width: int = 1) -> dict:
    if point:
        return {"type": "circle", "color": color, "radius": 5, "stroke": "#ffffff"}
    if fill:
        return {"type": "fill", "fillColor": color, "fillOpacity": 0.35,
                "color": color, "weight": width}
    return {"type": "line", "color": color, "weight": width}


def _scenario_style(period: int) -> dict:
    colors = {
        5: "#9ecae1",
        10: "#6baed6",
        20: "#4292c6",
        50: "#2171b5",
        100: "#084594",
    }
    return {"type": "fill", "fillColor": colors.get(period, "#3182bd"),
            "fillOpacity": 0.35, "color": colors.get(period, "#3182bd"),
            "weight": 0.5}
