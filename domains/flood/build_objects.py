from __future__ import annotations

import json
import math
import os
import re
import subprocess
from pathlib import Path
from typing import Any

import openpyxl

from .runtime.common import DATA_DIR, OBJECTS_DIR, PROJECT_DIR, SOURCES_DIR, rel
from .runtime.repository import object_library_path


SCENARIO_CODE_ALIASES = {
    "45050092shh0001": "45050092hsfx0001",
    "45050092shh0002": "45050092hsfx0002",
    "45050092shh0003": "45050092hsfx0003",
    "45050092shh0004": "45050092hsfx0004",
    "45050092shh0005": "45050092hsfx0005",
}


class FloodObjectBuilder:
    def __init__(self):
        self._cache: dict[str, list[dict]] = {}

    def build(self, object_type: str) -> list[dict]:
        if object_type in self._cache:
            return self._cache[object_type]
        method = getattr(self, f"_build_{object_type.lower()}", None)
        if not method:
            raise ValueError(f"unsupported object type: {object_type}")
        rows = method()
        self._cache[object_type] = rows
        return rows

    def _build_river(self) -> list[dict]:
        watershed = self.build("Watershed")[0]
        return [{
            "river_id": "shanhu",
            "name": "珊瑚河",
            "watershed_id": watershed["watershed_id"],
            "watershed_name": watershed["name"],
            "data_note": "OSM 未检索到命名为珊瑚河/Shanhu River 的 waterway 主体；河道绘制使用 Waterway 候选对象。",
        }]

    def _build_watershed(self) -> list[dict]:
        path = DATA_DIR / "1.流域边界/珊瑚河流域范围.shp"
        feature = _features(path)
        geo = _geometry_fields(feature[0]) if feature else {}
        return [{
            "watershed_id": "shanhu_watershed",
            "river_id": "shanhu",
            "name": "珊瑚河流域",
            "crs": "EPSG:2434",
            **geo,
            "data_path": rel(path),
        }]

    def _build_waterway(self) -> list[dict]:
        watershed = self.build("Watershed")[0]
        basin_geometry = json.loads(watershed["geometry"])
        bbox = _geometry_bbox(basin_geometry)
        source_path = _ensure_osm_waterways(bbox)
        source = json.loads(source_path.read_text(encoding="utf-8"))
        basin_polygon = _outer_ring(basin_geometry)
        rows = []
        for element in source.get("elements", []):
            geometry = element.get("geometry") or []
            if element.get("type") != "way" or len(geometry) < 2:
                continue
            coords = [(float(item["lon"]), float(item["lat"])) for item in geometry]
            length_m = _line_length_m(coords)
            inside_length_m = _inside_line_length_m(coords, basin_polygon)
            if inside_length_m <= 0:
                continue
            tags = element.get("tags") or {}
            osm_id = str(element["id"])
            rows.append({
                "waterway_id": f"osm_way_{osm_id}",
                "river_id": watershed["river_id"],
                "watershed_id": watershed["watershed_id"],
                "name": tags.get("name") or tags.get("name:zh") or "",
                "waterway_type": tags.get("waterway", ""),
                "osm_type": element["type"],
                "osm_id": osm_id,
                "source": "OpenStreetMap Overpass",
                "osm_source": tags.get("source", ""),
                "length_m": round(length_m, 2),
                "inside_basin_length_m": round(inside_length_m, 2),
                "inside_basin_ratio": round(inside_length_m / length_m, 4) if length_m else 0,
                "candidate_status": "osm_unnamed_candidate" if not (tags.get("name") or tags.get("name:zh")) else "osm_named_neighbor",
                "geometry_type": "LineString",
                "geometry": json.dumps({
                    "type": "LineString",
                    "coordinates": [[lon, lat] for lon, lat in coords],
                }, ensure_ascii=False),
                "data_path": rel(source_path),
            })
        rows.sort(key=lambda row: (row["inside_basin_length_m"], row["length_m"]), reverse=True)
        return rows

    def _build_county(self) -> list[dict]:
        path = DATA_DIR / "2.县界/珊瑚河县界.shp"
        rows = []
        for i, feature in enumerate(_features(path), 1):
            props = feature.get("properties") or {}
            adcode = _code(_first_non_empty(props, "adcode", "ADCODE")) or f"county_{i}"
            rows.append({
                "county_id": adcode,
                "river_id": "shanhu",
                "name": _first_non_empty(props, "Name", "name") or "钟山县",
                "adcode": adcode,
                **_geometry_fields(feature),
                "data_path": rel(path),
            })
        return rows

    def _build_town(self) -> list[dict]:
        path = DATA_DIR / "7.灾前&灾损数据/珊瑚河灾前信息.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0])
            rows.append({
                "town_id": _slug_id(name),
                "name": name,
                "county_id": "451122",
                "population": int(_float(row[2])),
                "gdp_10k_cny": _float(row[1]),
                "agriculture_10k_cny": _float(row[3]),
                "housing_property_10k_cny": _float(row[4]),
                "transport_industry_10k_cny": _float(row[5]),
                "commerce_10k_cny": _float(row[6]),
            })
        wb.close()
        return rows

    def _build_reservoir(self) -> list[dict]:
        path = DATA_DIR / "3.水利工程/水库.shp"
        osm_reservoirs = _load_osm_reservoir_candidates()
        esa_water = _load_esa_water_candidates()
        hydrolakes = _load_hydrolakes_candidates()
        rows = []
        used_osm_ids: set[str] = set()
        used_hylak_ids: set[int] = set()
        for i, feature in enumerate(_features(path), 1):
            row = _reservoir_record(i, feature, path)
            match = _match_osm_reservoir(row, osm_reservoirs, used_osm_ids)
            if match:
                _apply_osm_reservoir(row, match)
                used_osm_ids.add(match["osm_ref"])
            esa_match = _match_esa_water_reservoir(row, esa_water)
            if esa_match:
                _apply_esa_water_reservoir(row, esa_match)
            else:
                hydro_match = _match_hydrolakes_reservoir(row, hydrolakes, used_hylak_ids)
                if hydro_match:
                    _apply_hydrolakes_reservoir(row, hydro_match)
                    used_hylak_ids.add(hydro_match["hylak_id"])
            rows.append(row)
        return rows

    def _build_sluice(self) -> list[dict]:
        path = DATA_DIR / "3.水利工程/水闸.shp"
        osm_hydraulic = _load_osm_candidates("osm_hydraulic_structures_shanhu.json", _osm_hydraulic_kind)
        rows = []
        used_osm_ids: set[str] = set()
        for i, feature in enumerate(_features(path), 1):
            row = _sluice_record(i, feature, path)
            match = _match_nearest_osm(row, osm_hydraulic, used_osm_ids, max_distance_m=350)
            if match:
                _apply_osm_match(row, match, "local_inventory_with_osm_hydraulic_match")
                used_osm_ids.add(match["osm_ref"])
            rows.append(row)
        return rows

    def _build_bridge(self) -> list[dict]:
        path = DATA_DIR / "3.水利工程/桥梁.shp"
        osm_bridges = _load_osm_candidates("osm_bridges_shanhu.json", _osm_bridge_kind)
        rows = []
        used_osm_ids: set[str] = set()
        for i, feature in enumerate(_features(path), 1):
            row = _bridge_record(i, feature, path)
            match = _match_nearest_osm(row, osm_bridges, used_osm_ids, max_distance_m=160)
            if match:
                _apply_osm_match(row, match, "local_inventory_with_osm_bridge_match")
                used_osm_ids.add(match["osm_ref"])
            rows.append(row)
        return rows

    def _build_facility(self) -> list[dict]:
        specs = [
            ("hospital", DATA_DIR / "4.重要设施/医院.shp"),
            ("school", DATA_DIR / "4.重要设施/学校.shp"),
            ("government", DATA_DIR / "4.重要设施/政府.shp"),
        ]
        osm_facilities = _load_osm_candidates("osm_facilities_shanhu.json", _osm_facility_kind)
        rows = []
        used_osm_ids: set[str] = set()
        for facility_type, path in specs:
            for i, feature in enumerate(_features(path), 1):
                row = _facility_record(facility_type, i, feature, path)
                match = _match_facility_osm(row, osm_facilities, used_osm_ids)
                if match:
                    _apply_osm_match(row, match, "local_inventory_with_osm_facility_match")
                    used_osm_ids.add(match["osm_ref"])
                rows.append(row)
        return rows

    def _build_hydraulicstructure(self) -> list[dict]:
        specs = [
            ("levee", DATA_DIR / "3.水利工程/堤防.shp"),
            ("pump_station", DATA_DIR / "3.水利工程/泵站.shp"),
            ("spillway", DATA_DIR / "3.水利工程/溢流建筑物.shp"),
        ]
        osm_hydraulic = _load_osm_candidates("osm_hydraulic_structures_shanhu.json", _osm_hydraulic_kind)
        rows = []
        used_osm_ids: set[str] = set()
        for structure_type, path in specs:
            for i, feature in enumerate(_features(path), 1):
                row = _hydraulic_structure_record(structure_type, i, feature, path)
                match = _match_nearest_osm(row, osm_hydraulic, used_osm_ids, max_distance_m=350)
                if match:
                    _apply_osm_match(row, match, "local_inventory_with_osm_hydraulic_match")
                    used_osm_ids.add(match["osm_ref"])
                rows.append(row)
        return rows

    def _build_road(self) -> list[dict]:
        path = DATA_DIR / "5.路网/公路-线.shp"
        osm_roads = _load_osm_candidates("osm_roads_shanhu.json", _osm_road_kind)
        osm_by_id = {item["osm_id"]: item for item in osm_roads if item.get("osm_id")}
        rows = []
        for i, feature in enumerate(_features(path), 1):
            row = _road_record(i, feature, path)
            match = osm_by_id.get(str(row.get("road_id", "")))
            if match:
                _apply_osm_match(row, match, "local_inventory_with_osm_road_match")
                row["osm_match_status"] = "matched_by_osm_id"
            rows.append(row)
        return rows

    def _build_place(self) -> list[dict]:
        rows = []
        shelter_path = DATA_DIR / "8.避洪转移/安置点.shp"
        inplace_path = DATA_DIR / "8.避洪转移/就地单元.shp"
        for i, feature in enumerate(_features(shelter_path), 1):
            rows.append(_place_record("shelter", i, feature, shelter_path))
        for i, feature in enumerate(_features(inplace_path), 1):
            rows.append(_place_record("in_place", i, feature, inplace_path))
        return rows

    def _build_transfer(self) -> list[dict]:
        path = DATA_DIR / "8.避洪转移/转移单元.shp"
        transfer_by_id = {}
        for i, feature in enumerate(_features(path), 1):
            row = _transfer_record(i, feature, path)
            transfer_by_id[row["transfer_id"]] = row
        for route in self.build("Route"):
            transfer_id = route.get("transfer_id")
            if transfer_id in transfer_by_id:
                transfer_by_id[transfer_id]["route_id"] = route["route_id"]
                transfer_by_id[transfer_id]["place_id"] = route.get("place_id", "")
        return list(transfer_by_id.values())

    def _build_route(self) -> list[dict]:
        path = DATA_DIR / "8.避洪转移/转移路线.shp"
        return [_route_record(i, feature, path) for i, feature in enumerate(_features(path), 1)]

    def _build_scenario(self) -> list[dict]:
        path = DATA_DIR / "6.淹没图层/方案说明内容_河流.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, _, scenario_id, name, description, summary = row[:6]
            if not scenario_id:
                continue
            rows.append({
                "scenario_id": str(scenario_id),
                "river_id": "shanhu",
                "name": str(name or ""),
                "return_period_year": _return_period(name or summary or description),
                "summary": str(summary or ""),
                "description": str(description or ""),
                "data_path": rel(DATA_DIR / f"6.淹没图层/45050092_珊瑚河/{scenario_id}.shp"),
            })
        wb.close()
        return rows

    def _build_impact(self) -> list[dict]:
        path = DATA_DIR / "7.灾前&灾损数据/45050092_洪水影响分析与损失评估结果表.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            raw_id = str(row[0])
            scenario_id = SCENARIO_CODE_ALIASES.get(raw_id, raw_id)
            rows.append({
                "impact_id": f"impact_{scenario_id}",
                "scenario_id": scenario_id,
                "inundated_area_km2": _float(row[1]),
                "inundated_building_10k_m2": _float(row[2]),
                "inundated_farmland_ha": _float(row[3]),
                "inundated_road_km": _float(row[4]),
                "affected_population_10k": _float(row[5]),
                "affected_gdp_10k_cny": _float(row[6]),
                "direct_loss_10k_cny": _float(row[7]),
            })
        wb.close()
        return rows

    def _build_hydrology(self) -> list[dict]:
        path = DATA_DIR / "9.水文计算结果/珊瑚河水文分析结果.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        periods = [100, 50, 20, 10, 5]
        scenario_by_period = {
            row["return_period_year"]: row["scenario_id"]
            for row in self.build("Scenario")
            if row.get("return_period_year")
        }
        for r in range(4, 7):
            duration = _duration_hours(ws.cell(r, 1).value)
            for idx, period in enumerate(periods, start=5):
                rows.append({
                    "hydrology_id": f"rain_{duration:g}h_{period}a",
                    "river_id": "shanhu",
                    "scenario_id": scenario_by_period.get(period, ""),
                    "return_period_year": period,
                    "duration_h": duration,
                    "design_rainfall_mm": _float(ws.cell(r, idx).value),
                    "source": "珊瑚河水文分析结果.xlsx: 设计暴雨",
                })
        peak_values = [ws.cell(12, col).value for col in range(3, 8)]
        discharge_values = [ws.cell(13, col).value for col in range(3, 8)]
        for period, peak, discharge in zip(periods, peak_values, discharge_values):
            rows.append({
                "hydrology_id": f"flood_peak_{period}a",
                "river_id": "shanhu",
                "scenario_id": scenario_by_period.get(period, ""),
                "return_period_year": period,
                "peak_inflow_m3s": _float(peak),
                "max_discharge_m3s": _float(discharge),
                "source": "珊瑚河水文分析结果.xlsx: 设计洪水",
            })
        wb.close()
        return rows


def write_object_library(object_type: str, rows: list[dict]) -> Path:
    OBJECTS_DIR.mkdir(parents=True, exist_ok=True)
    path = object_library_path(object_type)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            handle.write("\n")
    return path


def _features(path: Path) -> list[dict]:
    result = subprocess.run(
        ["ogr2ogr", "-f", "GeoJSON", "/vsistdout/", "-t_srs", "EPSG:4326", str(path)],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0 or not result.stdout.strip():
        return []
    return json.loads(result.stdout).get("features", [])


def _ensure_osm_waterways(bbox: tuple[float, float, float, float]) -> Path:
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)
    path = SOURCES_DIR / "osm_waterways_shanhu.json"
    if path.exists():
        return path
    min_lon, min_lat, max_lon, max_lat = bbox
    query = (
        "[out:json][timeout:45];\n"
        f"way[\"waterway\"~\"^(river|stream)$\"]({min_lat:.6f},{min_lon:.6f},{max_lat:.6f},{max_lon:.6f});\n"
        "out geom tags;\n"
    )
    result = subprocess.run(
        [
            "curl",
            "-4",
            "--noproxy",
            "*",
            "--connect-timeout",
            "10",
            "--max-time",
            "90",
            "-sS",
            "-X",
            "POST",
            "https://overpass-api.de/api/interpreter",
            "-H",
            "User-Agent: flood-domain-builder/0.1 (local research)",
            "-H",
            "Accept: application/json",
            "--data-urlencode",
            f"data={query}",
        ],
        check=False,
        capture_output=True,
        text=True,
        env={
            key: value for key, value in os.environ.items()
            if key not in {"http_proxy", "https_proxy", "all_proxy", "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY"}
        },
    )
    if result.returncode != 0:
        raise RuntimeError(f"Overpass download failed: {result.stderr.strip()}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Overpass returned non-JSON response: {result.stdout[:500]}") from exc
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return path


def _ensure_osm_reservoirs() -> Path:
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)
    path = SOURCES_DIR / "osm_reservoirs_shanhu_region.json"
    if path.exists():
        return path
    reservoirs = _features(DATA_DIR / "3.水利工程/水库.shp")
    points = [_point_coords(feature) for feature in reservoirs]
    points = [(lon, lat) for lon, lat in points if lon and lat]
    if points:
        min_lon = min(lon for lon, _ in points) - 0.08
        min_lat = min(lat for _, lat in points) - 0.08
        max_lon = max(lon for lon, _ in points) + 0.08
        max_lat = max(lat for _, lat in points) + 0.08
    else:
        min_lon, min_lat, max_lon, max_lat = 111.0, 23.75, 111.9, 25.05
    query = (
        "[out:json][timeout:90];\n"
        "(\n"
        f"  nwr[\"natural\"=\"water\"][\"water\"=\"reservoir\"]({min_lat:.6f},{min_lon:.6f},{max_lat:.6f},{max_lon:.6f});\n"
        f"  nwr[\"landuse\"=\"reservoir\"]({min_lat:.6f},{min_lon:.6f},{max_lat:.6f},{max_lon:.6f});\n"
        f"  nwr[\"waterway\"=\"dam\"]({min_lat:.6f},{min_lon:.6f},{max_lat:.6f},{max_lon:.6f});\n"
        f"  nwr[\"reservoir_type\"]({min_lat:.6f},{min_lon:.6f},{max_lat:.6f},{max_lon:.6f});\n"
        ");\n"
        "out center geom tags;\n"
    )
    result = subprocess.run(
        [
            "curl",
            "-4",
            "--noproxy",
            "*",
            "--connect-timeout",
            "10",
            "--max-time",
            "140",
            "-sS",
            "-X",
            "POST",
            "https://overpass-api.de/api/interpreter",
            "-H",
            "User-Agent: flood-domain-builder/0.1 (local research)",
            "-H",
            "Accept: application/json",
            "--data-urlencode",
            f"data={query}",
        ],
        check=False,
        capture_output=True,
        text=True,
        env={
            key: value for key, value in os.environ.items()
            if key not in {"http_proxy", "https_proxy", "all_proxy", "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY"}
        },
    )
    if result.returncode != 0:
        raise RuntimeError(f"Overpass reservoir download failed: {result.stderr.strip()}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Overpass returned non-JSON response: {result.stdout[:500]}") from exc
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return path


def _load_osm_reservoir_candidates() -> list[dict]:
    source_path = _ensure_osm_reservoirs()
    source = json.loads(source_path.read_text(encoding="utf-8"))
    candidates = []
    for element in source.get("elements", []):
        tags = element.get("tags") or {}
        kind = _osm_reservoir_kind(tags)
        if not kind:
            continue
        center = _osm_center(element)
        if not center:
            continue
        geometry = _osm_geometry(element)
        candidates.append({
            "osm_ref": f"{element.get('type')}/{element.get('id')}",
            "osm_type": element.get("type", ""),
            "osm_id": str(element.get("id", "")),
            "name": tags.get("name") or tags.get("name:zh") or "",
            "kind": kind,
            "tags": tags,
            "center": center,
            "geometry": geometry,
            "source_path": source_path,
        })
    return candidates


def _load_osm_candidates(filename: str, kind_fn) -> list[dict]:
    source_path = SOURCES_DIR / filename
    if not source_path.exists():
        return []
    try:
        source = json.loads(source_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    candidates = []
    for element in source.get("elements", []):
        tags = element.get("tags") or {}
        kind = kind_fn(tags)
        if not kind:
            continue
        center = _osm_center(element)
        if not center:
            continue
        geometry = _osm_geometry(element)
        candidates.append({
            "osm_ref": f"{element.get('type')}/{element.get('id')}",
            "osm_type": element.get("type", ""),
            "osm_id": str(element.get("id", "")),
            "name": tags.get("name") or tags.get("name:zh") or "",
            "kind": kind,
            "tags": tags,
            "center": center,
            "geometry": geometry,
            "source_path": source_path,
        })
    return candidates


def _load_hydrolakes_candidates() -> list[dict]:
    source_path = SOURCES_DIR / "hydrolakes" / "hydrolakes_shanhu_bbox.geojson"
    if not source_path.exists():
        return []
    source = json.loads(source_path.read_text(encoding="utf-8"))
    candidates = []
    for feature in source.get("features", []):
        props = feature.get("properties") or {}
        geometry = feature.get("geometry") or {}
        hylak_id = int(_float(props.get("Hylak_id")))
        if not hylak_id or not geometry:
            continue
        candidates.append({
            "hylak_id": hylak_id,
            "name": props.get("Lake_name") or "",
            "lake_type": int(_float(props.get("Lake_type"))),
            "lake_area_km2": _float(props.get("Lake_area")),
            "geometry": geometry,
            "source_path": source_path,
        })
    return candidates


def _load_esa_water_candidates() -> list[dict]:
    source_paths = [
        SOURCES_DIR / "esa_worldcover" / "derived" / "esa_wc_water_N24E111_polygons.geojson",
        SOURCES_DIR / "esa_worldcover" / "derived" / "esa_wc_water_N21E111_polygons.geojson",
    ]
    candidates = []
    for source_path in source_paths:
        if not source_path.exists():
            continue
        source = json.loads(source_path.read_text(encoding="utf-8"))
        for index, feature in enumerate(source.get("features", []), 1):
            geometry = feature.get("geometry") or {}
            if not geometry:
                continue
            bbox = _geometry_bbox(geometry)
            area_m2 = _bbox_area_m2(bbox)
            if area_m2 < 400:
                continue
            candidates.append({
                "esa_ref": f"{source_path.stem}/{index}",
                "geometry": geometry,
                "bbox": bbox,
                "bbox_area_m2": area_m2,
                "source_path": source_path,
            })
    return candidates


def _reservoir_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_rid = _first_non_empty(props, "CD", "ennmcd", "水库_1", "水库编", "OBJECTID") or "reservoir"
    rid = f"{_code(raw_rid) or raw_rid}_{index}"
    return {
        "reservoir_id": str(rid),
        "name": _first_non_empty(props, "NAME", "ennm", "水库名") or f"水库{index}",
        "river_id": "shanhu" if _first_non_empty(props, "所在河", "RV_NAME") == "珊瑚河" else "",
        "county_id": _code(_first_non_empty(props, "行政区", "adcode")),
        "town_name": _first_non_empty(props, "乡（镇", "town_name") or "",
        "capacity_10k_m3": _float(_first_non_empty(props, "CAPACITY", "总库容")),
        "dam_top_elevation_m": _float(_first_non_empty(props, "MAX_H", "坝顶高")),
        "design_flood_standard_year": _float(_first_non_empty(props, "D_ST", "设计洪")),
        "check_flood_standard_year": _float(_first_non_empty(props, "CH_ST", "校核洪")),
        "max_discharge_m3s": _float(_first_non_empty(props, "V_MAXDIS", "最大泄")),
        "safety_status": str(_first_non_empty(props, "SAFETY", "审核标") or ""),
        "longitude": lng,
        "latitude": lat,
        "geometry_source": "local_inventory",
        "osm_match_status": "unmatched",
        "osm_ref": "",
        "osm_name": "",
        "osm_water_feature": "",
        "osm_distance_m": 0.0,
        "osm_tags": "",
        "osm_data_path": "",
        "external_geometry_source": "",
        "external_geometry_ref": "",
        "external_geometry_distance_m": 0.0,
        "external_geometry_confidence": "",
        "external_geometry_area_km2": 0.0,
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _match_osm_reservoir(row: dict, candidates: list[dict],
                         used_osm_ids: set[str]) -> dict | None:
    point = (row.get("longitude") or 0.0, row.get("latitude") or 0.0)
    name = row.get("name") or ""
    available = [item for item in candidates if item["osm_ref"] not in used_osm_ids]

    named = [
        item for item in available
        if name and item["name"] == name and item["kind"] == "reservoir_area"
        and _distance_m(point, item["center"]) <= 6000
    ]
    if named:
        return min(named, key=lambda item: _distance_m(point, item["center"]))

    nearby_areas = [
        item for item in available
        if item["kind"] == "reservoir_area"
        and _distance_m(point, item["center"]) <= 1200
    ]
    if nearby_areas:
        return min(nearby_areas, key=lambda item: _distance_m(point, item["center"]))

    nearby_dams = [
        item for item in available
        if item["kind"] == "dam"
        and _distance_m(point, item["center"]) <= 250
    ]
    if nearby_dams:
        return min(nearby_dams, key=lambda item: _distance_m(point, item["center"]))
    return None


def _apply_osm_reservoir(row: dict, match: dict):
    point = (row.get("longitude") or 0.0, row.get("latitude") or 0.0)
    distance = _distance_m(point, match["center"])
    tags = match["tags"]
    row.update({
        "osm_match_status": "matched_by_name" if match["name"] and match["name"] == row.get("name") else "matched_by_distance",
        "osm_ref": match["osm_ref"],
        "osm_name": match["name"],
        "osm_water_feature": match["kind"],
        "osm_distance_m": round(distance, 2),
        "osm_tags": json.dumps(_selected_osm_tags(tags), ensure_ascii=False, sort_keys=True),
        "osm_data_path": rel(match["source_path"]),
    })
    if match["kind"] == "reservoir_area" and match["geometry"]:
        row["geometry_source"] = "osm_reservoir_area"
        row["geometry_type"] = match["geometry"]["type"]
        row["geometry"] = json.dumps(match["geometry"], ensure_ascii=False)
    elif match["kind"] == "reservoir_area":
        row["geometry_source"] = "local_inventory_with_osm_reservoir_match"
    elif match["kind"] == "dam":
        row["geometry_source"] = "local_inventory_with_osm_dam_match"


def _match_hydrolakes_reservoir(row: dict, candidates: list[dict],
                                used_hylak_ids: set[int],
                                max_distance_m: float = 300) -> dict | None:
    point = _row_point(row)
    if not point:
        return None
    matches = []
    for item in candidates:
        if item["hylak_id"] in used_hylak_ids:
            continue
        distance, inside = _distance_to_geometry(point, item["geometry"])
        if inside or distance <= max_distance_m:
            matches.append({**item, "distance_m": distance, "inside": inside})
    if not matches:
        return None
    return min(matches, key=lambda item: (0 if item["inside"] else 1, item["distance_m"]))


def _apply_hydrolakes_reservoir(row: dict, match: dict):
    row.update({
        "geometry_source": "hydrolakes_reservoir_area",
        "geometry_type": match["geometry"]["type"],
        "geometry": json.dumps(match["geometry"], ensure_ascii=False),
        "external_geometry_source": "HydroLAKES",
        "external_geometry_ref": f"Hylak_id/{match['hylak_id']}",
        "external_geometry_distance_m": round(match["distance_m"], 2),
        "external_geometry_confidence": "high",
        "external_geometry_area_km2": match["lake_area_km2"],
    })


def _match_esa_water_reservoir(row: dict, candidates: list[dict],
                               max_distance_m: float = 100) -> dict | None:
    point = _row_point(row)
    if not point:
        return None
    matches = []
    for item in candidates:
        if not _point_near_bbox(point, item["bbox"], max_distance_m):
            continue
        distance, inside = _distance_to_geometry(point, item["geometry"])
        if inside or distance <= max_distance_m:
            matches.append({**item, "distance_m": distance, "inside": inside})
    if not matches:
        return None
    return min(matches, key=lambda item: (0 if item["inside"] else 1, item["distance_m"], -item["bbox_area_m2"]))


def _apply_esa_water_reservoir(row: dict, match: dict):
    row.update({
        "geometry_source": "esa_worldcover_water_area",
        "geometry_type": match["geometry"]["type"],
        "geometry": json.dumps(match["geometry"], ensure_ascii=False),
        "external_geometry_source": "ESA WorldCover",
        "external_geometry_ref": match["esa_ref"],
        "external_geometry_distance_m": round(match["distance_m"], 2),
        "external_geometry_confidence": "experimental",
        "external_geometry_area_km2": round(match["bbox_area_m2"] / 1_000_000, 6),
    })


def _osm_reservoir_kind(tags: dict) -> str:
    if tags.get("water") == "reservoir" or tags.get("landuse") == "reservoir" or "reservoir_type" in tags:
        return "reservoir_area"
    if tags.get("waterway") == "dam":
        return "dam"
    return ""


def _osm_center(element: dict) -> tuple[float, float] | None:
    center = element.get("center")
    if center:
        return float(center["lon"]), float(center["lat"])
    bounds = element.get("bounds")
    if bounds:
        lon = (float(bounds["minlon"]) + float(bounds["maxlon"])) / 2
        lat = (float(bounds["minlat"]) + float(bounds["maxlat"])) / 2
        return lon, lat
    if "lon" in element and "lat" in element:
        return float(element["lon"]), float(element["lat"])
    geometry = element.get("geometry") or []
    if not geometry:
        return None
    lon = sum(float(item["lon"]) for item in geometry) / len(geometry)
    lat = sum(float(item["lat"]) for item in geometry) / len(geometry)
    return lon, lat


def _osm_geometry(element: dict) -> dict:
    geometry = element.get("geometry") or []
    if not geometry:
        return {}
    coords = [[float(item["lon"]), float(item["lat"])] for item in geometry]
    tags = element.get("tags") or {}
    if _osm_reservoir_kind(tags) == "reservoir_area" and len(coords) >= 4:
        if coords[0] != coords[-1]:
            coords.append(coords[0])
        return {"type": "Polygon", "coordinates": [coords]}
    if len(coords) >= 2:
        return {"type": "LineString", "coordinates": coords}
    return {"type": "Point", "coordinates": coords[0]}


def _selected_osm_tags(tags: dict) -> dict:
    keys = [
        "name", "name:en", "name:zh", "name:zh-Latn-pinyin",
        "natural", "water", "landuse", "waterway", "reservoir_type",
        "highway", "ref", "oneway", "bridge", "tunnel", "layer", "surface",
        "amenity", "office", "government", "man_made", "pump",
        "code", "wikidata", "wikipedia", "source",
    ]
    return {key: tags[key] for key in keys if key in tags}


def _osm_road_kind(tags: dict) -> str:
    return tags.get("highway", "")


def _osm_bridge_kind(tags: dict) -> str:
    if tags.get("bridge") or tags.get("man_made") == "bridge":
        return "bridge"
    return ""


def _osm_facility_kind(tags: dict) -> str:
    amenity = tags.get("amenity", "")
    if amenity in {"school", "kindergarten", "college", "university"}:
        return "school"
    if amenity in {"hospital", "clinic", "doctors", "dentist", "pharmacy"}:
        return "hospital"
    if amenity in {"townhall", "public_building", "police", "fire_station"} or tags.get("office") == "government" or tags.get("government"):
        return "government"
    return ""


def _osm_hydraulic_kind(tags: dict) -> str:
    if tags.get("waterway") in {"dam", "weir", "sluice_gate", "lock_gate"}:
        return tags["waterway"]
    if tags.get("man_made") in {"dyke", "embankment", "pumping_station", "water_works"}:
        return tags["man_made"]
    return ""


def _osm_defaults() -> dict:
    return {
        "geometry_source": "local_inventory",
        "osm_match_status": "unmatched",
        "osm_ref": "",
        "osm_name": "",
        "osm_feature": "",
        "osm_distance_m": 0.0,
        "osm_tags": "",
        "osm_data_path": "",
    }


def _match_nearest_osm(row: dict, candidates: list[dict],
                       used_osm_ids: set[str], max_distance_m: float) -> dict | None:
    point = _row_point(row)
    if not point:
        return None
    available = [item for item in candidates if item["osm_ref"] not in used_osm_ids]
    named = [
        item for item in available
        if row.get("name") and item.get("name") == row.get("name")
        and _distance_m(point, item["center"]) <= max(max_distance_m, 1000)
    ]
    if named:
        return min(named, key=lambda item: _distance_m(point, item["center"]))
    nearby = [
        item for item in available
        if _distance_m(point, item["center"]) <= max_distance_m
    ]
    if nearby:
        return min(nearby, key=lambda item: _distance_m(point, item["center"]))
    return None


def _match_facility_osm(row: dict, candidates: list[dict],
                        used_osm_ids: set[str]) -> dict | None:
    wanted = row.get("facility_type", "")
    typed = [
        item for item in candidates
        if item["kind"] == wanted and item["osm_ref"] not in used_osm_ids
    ]
    point = _row_point(row)
    if not point:
        return None
    named = [
        item for item in typed
        if row.get("name") and item.get("name") == row.get("name")
        and _distance_m(point, item["center"]) <= 1200
    ]
    if named:
        return min(named, key=lambda item: _distance_m(point, item["center"]))
    nearby = [
        item for item in typed
        if _distance_m(point, item["center"]) <= 180
    ]
    if nearby:
        return min(nearby, key=lambda item: _distance_m(point, item["center"]))
    return None


def _apply_osm_match(row: dict, match: dict, geometry_source: str):
    point = _row_point(row)
    distance = _distance_m(point, match["center"]) if point else 0.0
    row.update({
        "geometry_source": geometry_source,
        "osm_match_status": "matched_by_name" if match["name"] and match["name"] == row.get("name") else "matched_by_distance",
        "osm_ref": match["osm_ref"],
        "osm_name": match["name"],
        "osm_feature": match["kind"],
        "osm_distance_m": round(distance, 2),
        "osm_tags": json.dumps(_selected_osm_tags(match["tags"]), ensure_ascii=False, sort_keys=True),
        "osm_data_path": rel(match["source_path"]),
    })


def _row_point(row: dict) -> tuple[float, float] | None:
    lon = row.get("longitude") or 0.0
    lat = row.get("latitude") or 0.0
    if lon and lat:
        return float(lon), float(lat)
    geom_text = row.get("geometry") or ""
    if not geom_text:
        return None
    try:
        geom = json.loads(geom_text)
    except json.JSONDecodeError:
        return None
    coords = list(_iter_coords(geom.get("coordinates") or []))
    if not coords:
        return None
    return (
        sum(lon for lon, _ in coords) / len(coords),
        sum(lat for _, lat in coords) / len(coords),
    )


def _sluice_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_sid = _first_non_empty(props, "CODE", "水闸编", "OBJECTID", "序号") or "sluice"
    sid = f"{_code(raw_sid) or raw_sid}_{index}"
    river_name = _first_non_empty(props, "RV_NAME", "所在河") or ""
    return {
        "sluice_id": str(sid),
        "name": _first_non_empty(props, "NAME", "水闸名") or f"水闸{index}",
        "river_id": "shanhu" if river_name == "珊瑚河" else "",
        "county_id": _code(_first_non_empty(props, "行政区", "adcode")),
        "town_name": _first_non_empty(props, "乡（镇") or "",
        "sluice_type": _first_non_empty(props, "TYPE", "水闸类") or "",
        "gate_count": int(_float(_first_non_empty(props, "闸孔数"))),
        "design_discharge_m3s": _float(_first_non_empty(props, "分_泄_", "D_MAXDIS")),
        "longitude": lng,
        "latitude": lat,
        **_osm_defaults(),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _bridge_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    name = _first_non_empty(props, "NAME", "ennm", "名称_nam")
    bridge_id = _first_non_empty(props, "CODE", "ennmcd", "Id") or f"bridge_{index}"
    populated = [value for value in props.values() if value not in (None, "", 0, 0.0, [])]
    return {
        "bridge_id": str(bridge_id) if str(bridge_id) != "0" else f"bridge_{index}",
        "name": name or f"桥梁{index}",
        "river_id": "shanhu" if _first_non_empty(props, "RV_NAME") == "珊瑚河" else "",
        "road_id": "",
        "length_m": _float(_first_non_empty(props, "LEN")),
        "width_m": _float(_first_non_empty(props, "WIDE")),
        "deck_elevation_m": _float(_first_non_empty(props, "EL", "B_EL")),
        "data_quality": "partial" if len(populated) > 3 else "missing",
        "longitude": lng,
        "latitude": lat,
        **_osm_defaults(),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _facility_record(facility_type: str, index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    fid = _first_non_empty(props, "id", "OBJECTID") or f"{facility_type}_{index}"
    return {
        "facility_id": str(fid),
        "name": _first_non_empty(props, "name") or f"{_facility_type_cn(facility_type)}{index}",
        "facility_type": facility_type,
        "subtype": _first_non_empty(props, "行业小", "行业中") or "",
        "address": _first_non_empty(props, "address") or "",
        "town_name": "",
        "county_id": _code(_first_non_empty(props, "adcode")),
        "longitude": lng or _float(_first_non_empty(props, "wgs84_经", "longitude")),
        "latitude": lat or _float(_first_non_empty(props, "wgs84_纬", "latitude")),
        **_osm_defaults(),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _hydraulic_structure_record(structure_type: str, index: int,
                                feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_sid = _first_non_empty(props, "ID", "CODE", "ennmcd", "OBJECTID") or structure_type
    sid = f"{_code(raw_sid) or raw_sid}_{index}"
    river_name = _first_non_empty(props, "RV_NAME", "LOC", "所在河") or ""
    return {
        "structure_id": str(sid),
        "name": _first_non_empty(props, "NAME", "ennm", "enname") or f"{_structure_type_cn(structure_type)}{index}",
        "structure_type": structure_type,
        "river_id": "shanhu" if river_name == "珊瑚河" else "",
        "river_name": river_name,
        "county_id": _code(_first_non_empty(props, "adcode", "行政区")),
        "location": _first_non_empty(props, "LOC") or "",
        "length_m": _float(_first_non_empty(props, "LEN", "Shape_Leng")),
        "elevation_m": _float(_first_non_empty(props, "EL", "MAX_H")),
        "flow_m3s": _float(_first_non_empty(props, "FLOW", "D_MAXDIS")),
        "longitude": lng,
        "latitude": lat,
        **_osm_defaults(),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _road_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    road_id = _first_non_empty(props, "osm_id", "OBJECTID") or f"road_{index}"
    return {
        "road_id": str(road_id),
        "name": _first_non_empty(props, "name") or _first_non_empty(props, "ref") or f"道路{index}",
        "ref": _first_non_empty(props, "ref") or "",
        "road_class": _first_non_empty(props, "fclass") or "",
        "one_way": _first_non_empty(props, "oneway") or "",
        "bridge_flag": str(_first_non_empty(props, "bridge") or "").upper() == "T",
        "tunnel_flag": str(_first_non_empty(props, "tunnel") or "").upper() == "T",
        "length_m": _float(_first_non_empty(props, "Shape_Leng")),
        "crs": "EPSG:4546",
        **_osm_defaults(),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _place_record(place_type: str, index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    raw_id = _first_non_empty(props, "ID", "OBJECTID") or f"{place_type}_{index}"
    return {
        "place_id": f"{place_type}_{_code(raw_id) or raw_id}",
        "name": _first_non_empty(props, "Name") or f"{_place_type_cn(place_type)}{index}",
        "place_type": place_type,
        "town_id": _code(_first_non_empty(props, "OwnerTown_")),
        "town_name": _first_non_empty(props, "OwnerTown") or "",
        "county_id": _code(_first_non_empty(props, "OwnerQX_Co")),
        "area_m2": _float(_first_non_empty(props, "Area")),
        "capacity_person": int(_float(_first_non_empty(props, "Vol", "Population"))),
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _transfer_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    lng, lat = _point_coords(feature)
    transfer_id = _code(_first_non_empty(props, "ID", "OBJECTID")) or f"transfer_{index}"
    return {
        "transfer_id": transfer_id,
        "name": _first_non_empty(props, "Name") or f"转移{index}",
        "population": int(_float(_first_non_empty(props, "Population"))),
        "town_id": _code(_first_non_empty(props, "OwnerTown_")),
        "town_name": _first_non_empty(props, "OwnerTown") or "",
        "county_id": _code(_first_non_empty(props, "OwnerQX_Co")),
        "arrive_time_window": _first_non_empty(props, "ArriveTime") or "",
        "route_id": "",
        "place_id": "",
        "longitude": lng,
        "latitude": lat,
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _route_record(index: int, feature: dict, path: Path) -> dict:
    props = feature.get("properties") or {}
    route_id = _code(_first_non_empty(props, "ID", "OBJECTID")) or f"route_{index}"
    transfer_id = _code(_first_non_empty(props, "waterID", "ID"))
    place_raw_id = _code(_first_non_empty(props, "safeID"))
    return {
        "route_id": route_id,
        "name": _first_non_empty(props, "TransferNa") or f"路线{index}",
        "route_type": "transfer",
        "road_detail": _first_non_empty(props, "RoadDetail") or "",
        "transfer_id": transfer_id,
        "place_id": f"shelter_{place_raw_id}" if place_raw_id else "",
        "population": int(_float(_first_non_empty(props, "Population"))),
        "length_m": _float(_first_non_empty(props, "Shape_Leng")),
        **_geometry_fields(feature),
        "data_path": rel(path),
    }


def _point_coords(feature: dict) -> tuple[float, float]:
    geom = feature.get("geometry") or {}
    coords = geom.get("coordinates") or []
    if geom.get("type") == "Point" and len(coords) >= 2:
        return float(coords[0]), float(coords[1])
    return 0.0, 0.0


def _geometry_bbox(geometry: dict) -> tuple[float, float, float, float]:
    coords = list(_iter_coords(geometry.get("coordinates") or []))
    lons = [item[0] for item in coords]
    lats = [item[1] for item in coords]
    return min(lons), min(lats), max(lons), max(lats)


def _bbox_area_m2(bbox: tuple[float, float, float, float]) -> float:
    min_lon, min_lat, max_lon, max_lat = bbox
    lat = (min_lat + max_lat) / 2
    width = abs(max_lon - min_lon) * 111320 * math.cos(math.radians(lat))
    height = abs(max_lat - min_lat) * 111320
    return width * height


def _point_near_bbox(point: tuple[float, float],
                     bbox: tuple[float, float, float, float],
                     max_distance_m: float) -> bool:
    lon, lat = point
    min_lon, min_lat, max_lon, max_lat = bbox
    pad_lat = max_distance_m / 111320
    pad_lon = max_distance_m / (111320 * math.cos(math.radians(lat)))
    return (
        min_lon - pad_lon <= lon <= max_lon + pad_lon
        and min_lat - pad_lat <= lat <= max_lat + pad_lat
    )


def _iter_coords(value):
    if not value:
        return
    if isinstance(value[0], (int, float)):
        yield float(value[0]), float(value[1])
        return
    for item in value:
        yield from _iter_coords(item)


def _outer_ring(geometry: dict) -> list[tuple[float, float]]:
    coords = geometry.get("coordinates") or []
    if geometry.get("type") == "Polygon" and coords:
        return [(float(lon), float(lat)) for lon, lat in coords[0]]
    if geometry.get("type") == "MultiPolygon" and coords and coords[0]:
        return [(float(lon), float(lat)) for lon, lat in coords[0][0]]
    return []


def _point_in_polygon(point: tuple[float, float], polygon: list[tuple[float, float]]) -> bool:
    x, y = point
    inside = False
    j = len(polygon) - 1
    for i, current in enumerate(polygon):
        xi, yi = current
        xj, yj = polygon[j]
        if (yi > y) != (yj > y):
            x_intersect = (xj - xi) * (y - yi) / ((yj - yi) or 1e-30) + xi
            if x < x_intersect:
                inside = not inside
        j = i
    return inside


def _distance_to_geometry(point: tuple[float, float], geometry: dict) -> tuple[float, bool]:
    rings = _geometry_rings(geometry)
    if not rings:
        return math.inf, False
    inside = any(_point_in_polygon(point, ring) for ring in rings if len(ring) >= 3)
    if inside:
        return 0.0, True
    distance = math.inf
    for ring in rings:
        for index in range(1, len(ring)):
            distance = min(distance, _distance_point_to_segment_m(point, ring[index - 1], ring[index]))
    return distance, False


def _geometry_rings(geometry: dict) -> list[list[tuple[float, float]]]:
    coords = geometry.get("coordinates") or []
    if geometry.get("type") == "Polygon":
        return [[(float(lon), float(lat)) for lon, lat in ring] for ring in coords]
    if geometry.get("type") == "MultiPolygon":
        return [
            [(float(lon), float(lat)) for lon, lat in ring]
            for polygon in coords
            for ring in polygon
        ]
    return []


def _distance_point_to_segment_m(point: tuple[float, float],
                                 start: tuple[float, float],
                                 end: tuple[float, float]) -> float:
    lon, lat = point
    lat0 = math.radians(lat)

    def project(item: tuple[float, float]) -> tuple[float, float]:
        item_lon, item_lat = item
        return (
            math.radians(item_lon - lon) * 6371000 * math.cos(lat0),
            math.radians(item_lat - lat) * 6371000,
        )

    px, py = 0.0, 0.0
    ax, ay = project(start)
    bx, by = project(end)
    dx = bx - ax
    dy = by - ay
    if dx == 0 and dy == 0:
        return math.hypot(px - ax, py - ay)
    t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / (dx * dx + dy * dy)))
    return math.hypot(px - (ax + t * dx), py - (ay + t * dy))


def _line_length_m(coords: list[tuple[float, float]]) -> float:
    return sum(_distance_m(coords[index - 1], coords[index]) for index in range(1, len(coords)))


def _inside_line_length_m(coords: list[tuple[float, float]], polygon: list[tuple[float, float]]) -> float:
    total = 0.0
    for index in range(1, len(coords)):
        start = coords[index - 1]
        end = coords[index]
        if _point_in_polygon(start, polygon) or _point_in_polygon(end, polygon):
            total += _distance_m(start, end)
    return total


def _distance_m(start: tuple[float, float], end: tuple[float, float]) -> float:
    lon1, lat1 = start
    lon2, lat2 = end
    radius = 6371000
    x = math.radians(lon2 - lon1) * math.cos(math.radians((lat1 + lat2) / 2))
    y = math.radians(lat2 - lat1)
    return radius * math.hypot(x, y)


def _geometry_fields(feature: dict) -> dict:
    geom = feature.get("geometry") or {}
    return {
        "geometry_type": geom.get("type", ""),
        "geometry": json.dumps(geom, ensure_ascii=False) if geom else "",
    }


def _first_non_empty(values: dict, *keys: str) -> Any:
    for key in keys:
        value = values.get(key)
        if value not in (None, "", [], {}, "null"):
            return value
    return None


def _float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _return_period(value: Any) -> int:
    match = re.search(r"(\d+)\s*(?:a|年)", str(value))
    return int(match.group(1)) if match else 0


def _duration_hours(value: Any) -> float:
    match = re.search(r"(\d+(?:\.\d+)?)", str(value))
    return float(match.group(1)) if match else 0


def _slug_id(value: str) -> str:
    mapping = {
        "回龙镇": "451122104",
        "石龙镇": "451122105",
        "凤翔镇": "451122106",
        "珊瑚镇": "451122107",
        "同古镇": "451122108",
        "清塘镇": "451122111",
    }
    return mapping.get(value, value)


def _facility_type_cn(value: str) -> str:
    return {"hospital": "医院", "school": "学校", "government": "政府"}.get(value, "设施")


def _place_type_cn(value: str) -> str:
    return {"shelter": "安置点", "in_place": "就地安置点"}.get(value, "地点")


def _structure_type_cn(value: str) -> str:
    return {
        "levee": "堤防",
        "pump_station": "泵站",
        "spillway": "溢流建筑物",
    }.get(value, "水利工程")


def _code(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    return text[:-2] if text.endswith(".0") else text
